#!/usr/bin/env python3
"""
Step 1: Book-level Metadata Acquisition via Wikidata
=====================================================
Automatically query Wikidata for bibliographic metadata of historical Chinese
texts and produce a standardised Excel spreadsheet.

Usage:
    python step1_metadata.py booklist.csv -o output.xlsx

Dependencies:
    pip install requests openpyxl
"""

import csv
import re
import sys
import json
import time
import argparse
import logging
from pathlib import Path
from typing import Optional

try:
    import requests
except ImportError:
    sys.exit("[ERROR] Missing 'requests' library. Run: pip install requests")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("[ERROR] Missing 'openpyxl' library. Run: pip install openpyxl")

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
WIKIDATA_API = "https://www.wikidata.org/w/api.php"
SEARCH_LANGS = ["zh", "zh-hans", "zh-hant", "en"]
REQUEST_DELAY = 0.5  # seconds between API calls to avoid rate-limiting

# Wikidata property IDs
PROP_AUTHOR = "P50"
PROP_PUB_DATE = "P577"
PROP_COUNTRY = "P495"
PROP_OFFICIAL_NAME = "P1449"
PROP_INSTANCE_OF = "P31"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Wikidata API client
# ---------------------------------------------------------------------------
class WikidataClient:
    """Lightweight wrapper around the Wikidata REST API."""

    def __init__(self, delay: float = REQUEST_DELAY):
        self.session = requests.Session()
        self.session.headers.update({
            "User-Agent": "ResearchMetadataBot/1.0 (Academic research project)"
        })
        self.delay = delay
        self._cache: dict = {}

    def _get(self, params: dict) -> dict:
        """Send a GET request and return parsed JSON."""
        params["format"] = "json"
        time.sleep(self.delay)
        try:
            resp = self.session.get(WIKIDATA_API, params=params, timeout=15)
            resp.raise_for_status()
            return resp.json()
        except requests.RequestException as e:
            log.warning("API request failed: %s", e)
            return {}

    def search_entity(self, query: str, language: str = "zh",
                      entity_type: str = "item", limit: int = 5) -> list[dict]:
        """Search for Wikidata entities by text query."""
        data = self._get({
            "action": "wbsearchentities",
            "search": query,
            "language": language,
            "type": entity_type,
            "limit": limit,
        })
        return data.get("search", [])

    def get_entity(self, qid: str) -> Optional[dict]:
        """Fetch full entity data by QID. Results are cached in memory."""
        if qid in self._cache:
            return self._cache[qid]

        data = self._get({
            "action": "wbgetentities",
            "ids": qid,
            "languages": "|".join(SEARCH_LANGS + ["en"]),
            "props": "labels|descriptions|claims|aliases",
        })
        entities = data.get("entities", {})
        entity = entities.get(qid)
        if entity:
            self._cache[qid] = entity
        return entity

    def get_entity_label(self, qid: str, lang: str = "zh-hans") -> str:
        """Return the label (display name) for an entity in the given language."""
        entity = self.get_entity(qid)
        if not entity:
            return ""
        labels = entity.get("labels", {})
        for try_lang in [lang, "zh", "zh-hans", "zh-hant", "en"]:
            if try_lang in labels:
                return labels[try_lang]["value"]
        return ""


# ---------------------------------------------------------------------------
# Metadata extractor
# ---------------------------------------------------------------------------
class MetadataExtractor:
    """Extract task-brief fields from a Wikidata entity."""

    def __init__(self, client: WikidataClient):
        self.client = client

    def extract_all(self, qid: str) -> dict:
        """Return a dict of all required metadata fields for a given QID."""
        entity = self.client.get_entity(qid)
        if not entity:
            return self._empty_result(qid)

        return {
            "wikiID": qid,
            "wikidata_url": f"https://www.wikidata.org/wiki/{qid}",
            "title_sCN": self._get_label(entity, "zh-hans"),
            "title_EN": self._get_label(entity, "en"),
            "title_default": self._get_official_name(entity),
            "book_description": self._get_description(entity, "en"),
            "book_description_zh": self._get_description(entity, "zh-hans"),
            "author_sCN": self._get_author(entity, "zh-hans"),
            "author_EN": self._get_author(entity, "en"),
            "author_qid": self._get_author_qid(entity),
            "pub_date": self._get_pub_date(entity),
            "origin_country": self._get_country(entity),
        }

    def _empty_result(self, qid: str = "N/A") -> dict:
        return {
            "wikiID": qid,
            "wikidata_url": "",
            "title_sCN": "", "title_EN": "", "title_default": "",
            "book_description": "", "book_description_zh": "",
            "author_sCN": "", "author_EN": "", "author_qid": "",
            "pub_date": "", "origin_country": "",
        }

    def _get_label(self, entity: dict, lang: str) -> str:
        labels = entity.get("labels", {})
        fallback_order = {
            "zh-hans": ["zh-hans", "zh", "zh-hant", "zh-cn"],
            "zh-hant": ["zh-hant", "zh", "zh-hans"],
            "en": ["en"],
            "zh": ["zh", "zh-hans", "zh-hant"],
        }
        for try_lang in fallback_order.get(lang, [lang]):
            if try_lang in labels:
                return labels[try_lang]["value"]
        return ""

    def _get_description(self, entity: dict, lang: str) -> str:
        descs = entity.get("descriptions", {})
        fallback_order = {
            "zh-hans": ["zh-hans", "zh", "zh-hant", "en"],
            "en": ["en", "zh-hans", "zh"],
        }
        for try_lang in fallback_order.get(lang, [lang]):
            if try_lang in descs:
                return descs[try_lang]["value"]
        return ""

    def _get_claim_values(self, entity: dict, prop: str) -> list[dict]:
        claims = entity.get("claims", {})
        return claims.get(prop, [])

    def _get_official_name(self, entity: dict) -> str:
        """P1449 official name -- typically the romanised / pinyin title."""
        for claim in self._get_claim_values(entity, PROP_OFFICIAL_NAME):
            mainsnak = claim.get("mainsnak", {})
            datavalue = mainsnak.get("datavalue", {})
            if datavalue.get("type") == "monolingualtext":
                return datavalue["value"]["text"]
        # Fallback: look for pinyin-like strings in aliases
        aliases = entity.get("aliases", {})
        for lang in ["en", "zh-latn"]:
            if lang in aliases:
                for alias in aliases[lang]:
                    val = alias.get("value", "")
                    if val and re.search(r"[a-zA-Z]", val):
                        return val
        return ""

    def _get_author_qid(self, entity: dict) -> str:
        for claim in self._get_claim_values(entity, PROP_AUTHOR):
            mainsnak = claim.get("mainsnak", {})
            datavalue = mainsnak.get("datavalue", {})
            if datavalue.get("type") == "wikibase-entityid":
                return datavalue["value"]["id"]
        return ""

    def _get_author(self, entity: dict, lang: str) -> str:
        qid = self._get_author_qid(entity)
        if qid:
            return self.client.get_entity_label(qid, lang)
        return ""

    def _get_pub_date(self, entity: dict) -> str:
        for claim in self._get_claim_values(entity, PROP_PUB_DATE):
            mainsnak = claim.get("mainsnak", {})
            datavalue = mainsnak.get("datavalue", {})
            if datavalue.get("type") == "time":
                time_val = datavalue["value"]
                time_str = time_val.get("time", "")
                precision = time_val.get("precision", 9)
                return self._format_time(time_str, precision)
        return ""

    def _format_time(self, time_str: str, precision: int) -> str:
        """Convert Wikidata time value to a human-readable string."""
        match = re.match(r"[+\-](\d+)-(\d+)-(\d+)", time_str)
        if not match:
            return time_str
        year, month, day = match.groups()
        year = int(year)
        if precision <= 7:
            century = (year // 100) + 1
            return f"{century}th century"
        elif precision == 8:
            return f"{year}s"
        elif precision == 9:
            return str(year)
        elif precision == 10:
            return f"{year}-{month}"
        else:
            return f"{year}-{month}-{day}"

    def _get_country(self, entity: dict) -> str:
        for claim in self._get_claim_values(entity, PROP_COUNTRY):
            mainsnak = claim.get("mainsnak", {})
            datavalue = mainsnak.get("datavalue", {})
            if datavalue.get("type") == "wikibase-entityid":
                country_qid = datavalue["value"]["id"]
                label = self.client.get_entity_label(country_qid, "zh-hans")
                if not label:
                    label = self.client.get_entity_label(country_qid, "en")
                return label
        return ""


# ---------------------------------------------------------------------------
# Search and matching logic
# ---------------------------------------------------------------------------
def clean_title(raw_title: str) -> str:
    """Strip edition info, parenthetical notes, etc. from a raw book title."""
    cleaned = re.sub(r"\s*[(\uff08].*?[)\uff09]", "", raw_title)
    cleaned = re.sub(r"\s*(Low-res|v\d+).*", "", cleaned, flags=re.IGNORECASE)
    return cleaned.strip()


def find_best_match(client: WikidataClient, title_cn: str,
                    edition_info: str = "") -> Optional[str]:
    """
    Search Wikidata for the best-matching QID given a Chinese book title.

    Strategies tried in order:
      1. Exact Chinese title
      2. Title + qualifier keywords
      3. Author name extracted from edition_info
    """
    cleaned = clean_title(title_cn)

    # Strategy 1: direct search with the Chinese title
    for lang in ["zh", "zh-hans", "en"]:
        results = client.search_entity(cleaned, language=lang, limit=8)
        qid = _pick_best(client, results, cleaned)
        if qid:
            return qid

    # Strategy 2: add qualifier keywords
    for suffix in ["book", "ancient text", "martial arts manual"]:
        results = client.search_entity(f"{cleaned} {suffix}", language="zh", limit=5)
        qid = _pick_best(client, results, cleaned)
        if qid:
            return qid

    # Strategy 3: extract author name from edition_info and search together
    author_match = re.search(r"[.\u00b7]?\s*(\w{2,4})\s*[\u64b0\u8457\u7f16\u8f91\u7e82]",
                             edition_info)
    if author_match:
        author_name = author_match.group(1)
        results = client.search_entity(
            f"{cleaned} {author_name}", language="zh", limit=5
        )
        qid = _pick_best(client, results, cleaned)
        if qid:
            return qid

    return None


def _pick_best(client: WikidataClient, results: list[dict],
               target_title: str) -> Optional[str]:
    """Select the best matching entity from a list of search results."""
    if not results:
        return None

    target_chars = set(target_title)

    for result in results:
        qid = result.get("id", "")
        label = result.get("label", "")
        description = result.get("description", "")

        # Skip obviously irrelevant results (films, people, ships, etc.)
        skip_keywords = ["film", "movie", "ship", "person",
                         "disambiguation"]
        if any(kw in description.lower() for kw in skip_keywords):
            continue

        # Check character overlap between label and target title
        label_chars = set(label)
        overlap = len(target_chars & label_chars)
        if overlap >= len(target_chars) * 0.5:
            entity = client.get_entity(qid)
            if entity and _is_likely_book(entity):
                return qid

    return None


def _is_likely_book(entity: dict) -> bool:
    """Heuristic check: does this entity look like a book or written work?"""
    claims = entity.get("claims", {})

    # Check P31 (instance of) against known book-type QIDs
    instance_claims = claims.get(PROP_INSTANCE_OF, [])
    book_types = {
        "Q571",      # book
        "Q47461344", # written work
        "Q7725634",  # literary work
        "Q49848",    # document
        "Q234460",   # text
        "Q1266946",  # treatise
        "Q5185279",  # manual
        "Q55439712", # military manual
    }
    for claim in instance_claims:
        mainsnak = claim.get("mainsnak", {})
        datavalue = mainsnak.get("datavalue", {})
        if datavalue.get("type") == "wikibase-entityid":
            type_qid = datavalue["value"]["id"]
            if type_qid in book_types:
                return True

    # Has author (P50) -> likely a book
    if PROP_AUTHOR in claims:
        return True

    # Has publication date (P577) -> likely a book
    if PROP_PUB_DATE in claims:
        return True

    # Description contains book-related keywords
    for lang in ["en", "zh-hans", "zh"]:
        desc = entity.get("descriptions", {}).get(lang, {}).get("value", "")
        book_words = ["book", "manual", "treatise", "text", "classic",
                      "encyclopedia"]
        if any(w in desc.lower() for w in book_words):
            return True

    return False


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------
def write_excel(books: list[dict], output_path: str):
    """Write metadata to a formatted Excel workbook."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Step1_Book_Metadata"

    # Styles
    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_font = Font(name="Arial", size=10)
    data_align = Alignment(vertical="center", wrap_text=True)
    yellow_fill = PatternFill("solid", fgColor="FFF2CC")
    green_fill = PatternFill("solid", fgColor="E2EFDA")
    red_font = Font(name="Arial", size=10, color="FF0000", italic=True)
    link_font = Font(name="Arial", size=10, color="0563C1", underline="single")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # Headers
    headers = [
        "Seq", "bookID", "wikiID", "title_sCN", "title_EN",
        "title_default", "book_description", "author_sCN", "author_EN",
        "pub_date", "origin_country", "cnt_volumes",
        "Edition/Source", "Status", "Wikidata URL",
    ]
    col_widths = [5, 12, 14, 16, 42, 25, 50, 14, 28, 22, 18, 12, 35, 30, 40]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows
    for i, book in enumerate(books):
        row = i + 2
        values = [
            book.get("seq", ""),
            book.get("bookID", ""),
            book.get("wikiID", "N/A"),
            book.get("title_sCN", ""),
            book.get("title_EN", ""),
            book.get("title_default", ""),
            book.get("book_description", ""),
            book.get("author_sCN", ""),
            book.get("author_EN", ""),
            book.get("pub_date", ""),
            book.get("origin_country", ""),
            book.get("cnt_volumes", 0),
            book.get("edition_source", ""),
            book.get("status", ""),
            book.get("wikidata_url", ""),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border

        # Colour the status cell based on match result
        status_cell = ws.cell(row=row, column=14)
        status_text = str(book.get("status", ""))
        if "AUTO_MATCHED" in status_text:
            status_cell.fill = green_fill
        elif "MANUAL_REQUIRED" in status_text:
            status_cell.fill = yellow_fill
            status_cell.font = red_font

        # Make the Wikidata URL column a clickable hyperlink
        url_cell = ws.cell(row=row, column=15)
        url = book.get("wikidata_url", "")
        if url:
            url_cell.font = link_font
            url_cell.hyperlink = url

    # Column widths
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze header row and enable auto-filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(books) + 1}"

    # --- Summary sheet ---
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Step 1 Metadata Acquisition - Summary"
    ws2["A1"].font = Font(name="Arial", bold=True, size=14)

    matched = sum(1 for b in books if "AUTO_MATCHED" in str(b.get("status", "")))
    manual = sum(1 for b in books if "MANUAL_REQUIRED" in str(b.get("status", "")))

    summary_data = [
        ("Total entries", len(books)),
        ("Wikidata auto-matched", matched),
        ("Needs manual review", manual),
        ("", ""),
        ("Next Steps:", ""),
        ("1.", "Review rows marked MANUAL_REQUIRED and fill in missing fields"),
        ("2.", "Fill in cnt_volumes (volume count) -- not available via Wikidata"),
        ("3.", "Verify title_EN translations match your project conventions"),
        ("4.", "For entries with wikiID = N/A, confirm with supervisor whether "
               "to create new Wikidata entries or leave as N/A"),
    ]
    for idx, (k, v) in enumerate(summary_data, 3):
        ws2.cell(row=idx, column=1, value=k).font = Font(
            name="Arial", size=10, bold=(idx < 7)
        )
        ws2.cell(row=idx, column=2, value=v).font = Font(name="Arial", size=10)
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 70

    wb.save(output_path)
    log.info("Excel saved: %s", output_path)


# ---------------------------------------------------------------------------
# Main pipeline
# ---------------------------------------------------------------------------
def read_booklist(csv_path: str) -> list[dict]:
    """Read the book list from a CSV file."""
    books = []
    with open(csv_path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            books.append({
                "seq": row.get("Sequence", "").strip(),
                "bookID": row.get("bookID", "").strip(),
                "raw_title": row.get("Reference_Title", "").strip(),
                "edition_source": row.get("Edition_Source", "").strip(),
            })
    log.info("Loaded %d book entries from CSV", len(books))
    return books


def process_books(books: list[dict], known_qids: dict = None) -> list[dict]:
    """
    Core processing loop: query Wikidata for each book and extract metadata.

    Args:
        books: list of dicts read from the CSV
        known_qids: optional bookID -> QID mapping to skip searching
    """
    if known_qids is None:
        known_qids = {}

    client = WikidataClient(delay=REQUEST_DELAY)
    extractor = MetadataExtractor(client)
    results = []

    for i, book in enumerate(books):
        bid = book["bookID"]
        title = book["raw_title"]
        edition = book["edition_source"]
        log.info("[%d/%d] Processing: %s - %s", i + 1, len(books), bid, title)

        # 1) Check for a pre-assigned QID
        qid = known_qids.get(bid)

        # 2) Otherwise, search Wikidata
        if not qid:
            cleaned_title = clean_title(title)
            log.info("  Searching Wikidata for: '%s'", cleaned_title)
            qid = find_best_match(client, title, edition)

        # 3) Extract metadata
        if qid:
            log.info("  Matched: %s", qid)
            metadata = extractor.extract_all(qid)
            status = f"AUTO_MATCHED ({qid})"
        else:
            log.info("  No match found")
            metadata = extractor._empty_result("N/A")
            status = "MANUAL_REQUIRED"

        # 4) Merge into result record
        result = {
            "seq": book["seq"],
            "bookID": bid,
            "edition_source": edition,
            "cnt_volumes": 0,  # must be filled manually
            "status": status,
            **metadata,
        }

        # Fall back to raw title if Wikidata returned nothing for title_sCN
        if not result["title_sCN"]:
            result["title_sCN"] = clean_title(title)

        results.append(result)

    return results


def main():
    parser = argparse.ArgumentParser(
        description="Step 1: Batch-query Wikidata for historical text metadata",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python step1_metadata.py booklist.csv
  python step1_metadata.py booklist.csv -o my_metadata.xlsx
  python step1_metadata.py booklist.csv --known-qids known_qids.json
  python step1_metadata.py booklist.csv --delay 1.0

CSV format (UTF-8):
  Sequence,bookID,Reference_Title,Edition_Source
  01,jx_mg,New Treatise on Military Efficiency (18 vols),Ming Qi Jiguang
        """
    )
    parser.add_argument("csv_file", help="Path to input CSV book list")
    parser.add_argument("-o", "--output", default="Step1_Book_Metadata.xlsx",
                        help="Output Excel path (default: Step1_Book_Metadata.xlsx)")
    parser.add_argument("--delay", type=float, default=0.5,
                        help="Seconds between API requests (default: 0.5)")
    parser.add_argument("--known-qids", type=str, default=None,
                        help="JSON file mapping bookID to known Wikidata QIDs")

    args = parser.parse_args()

    global REQUEST_DELAY
    REQUEST_DELAY = args.delay

    # Load pre-assigned QIDs if provided
    known_qids = {}
    if args.known_qids:
        with open(args.known_qids, "r", encoding="utf-8") as f:
            known_qids = json.load(f)
        log.info("Loaded %d pre-assigned QIDs", len(known_qids))

    log.info("=" * 60)
    log.info("Step 1: Book-level Metadata Acquisition")
    log.info("=" * 60)

    books = read_booklist(args.csv_file)
    if not books:
        sys.exit("[ERROR] CSV file is empty or has incorrect format")

    results = process_books(books, known_qids)
    write_excel(results, args.output)

    # Print summary
    matched = sum(1 for r in results if "AUTO_MATCHED" in r["status"])
    manual = sum(1 for r in results if "MANUAL_REQUIRED" in r["status"])
    log.info("=" * 60)
    log.info("Done. Total: %d | Auto-matched: %d | Manual needed: %d",
             len(results), matched, manual)
    log.info("Output file: %s", args.output)
    log.info("=" * 60)


if __name__ == "__main__":
    main()
